import os
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import subprocess

total_files_count = 0
processed_files_count = 0

def count_pst_files(folder):
    admin_folder = os.path.join(folder, "admin")
    pst_files = [file for file in os.listdir(admin_folder) if file.endswith('.pst')]
    return len(pst_files)

def compare_folders(base_folder):
    try:
        def get_files_names(folder):
            file_names = []
            for root, _, files in os.walk(folder):
                for file in files:
                    file_names.append(os.path.splitext(file)[0])  # Excluding file extensions
            return file_names

        base_folder = base_folder.replace("\\", "/")
        source_folder = os.path.join(base_folder, "source").replace("\\", "/")
        source_cc_folder = os.path.join(base_folder, "source_cc").replace("\\", "/")
        bilinguals_folder = os.path.join(base_folder, "bilinguals").replace("\\", "/")
        sent_folder = os.path.join(base_folder, "sent").replace("\\", "/")
        quote_folder = os.path.join(base_folder, "admin\\Quote").replace("\\", "/")
        invoice_folder = os.path.join(base_folder, "admin\\_Invoice").replace("\\", "/")
        reference_folder = os.path.join(base_folder, "reference").replace("\\", "/")
        mqarch_folder = "P:/_MMQGarbage/Archive"  # MQArch folder path

        source_files = get_files_names(source_folder)
        source_cc_files = get_files_names(source_cc_folder)
        bilinguals_files = get_files_names(bilinguals_folder)
        sent_files = get_files_names(sent_folder)
        quote_files = get_files_names(quote_folder)
        invoice_files = get_files_names(invoice_folder)
        reference_files = get_files_names(reference_folder)

        # Count PST files in admin folder
        pst_count = count_pst_files(base_folder)

        # Initialize an empty dictionary to store file existence
        file_existence = {}

        # Check files in the source folder
        for filename in source_files:
            file_existence[filename] = [
                '✔' if filename in bilinguals_files else '✘',
                '✔' if filename in sent_files else '✘'
            ]

        # Check files in the source_cc folder
        for filename in source_cc_files:
            if filename not in file_existence:
                file_existence[filename] = [
                    '✔' if filename in bilinguals_files else '✘',
                    '✔' if filename in sent_files else '✘'
                ]

        # Check files in the quote folder
        for filename in quote_files:
            if filename not in file_existence:
                file_existence[filename] = [
                    '✔' if filename in invoice_files else '✘',
                    '✔' if filename in reference_files else '✘'
                ]

        # Check files in the invoice folder
        for filename in invoice_files:
            if filename not in file_existence:
                file_existence[filename] = [
                    '✔' if filename in quote_files else '✘',
                    '✔' if filename in reference_files else '✘'
                ]

        # Check files in the reference folder
        for filename in reference_files:
            if filename not in file_existence:
                file_existence[filename] = [
                    '✔' if filename in quote_files else '✘',
                    '✔' if filename in invoice_files else '✘'
                ]

        # Add PST count to the dictionary
        file_existence['PST Count'] = pst_count

        return file_existence
    except Exception as e:
        print(f"An error occurred: {e}")
        return {}  # Return an empty dictionary in case of any error

def export_to_xlsx():
    # Suggest a name for the Excel file incorporating the name of the chosen folder
    default_file_name = f"AUDIT_{os.path.basename(os.path.normpath(base_folder))}.xlsx"

    xlsx_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                                                  initialfile=default_file_name)
    if xlsx_file_path:
        file_existence = compare_folders(base_folder)
        workbook = openpyxl.Workbook()

        # First workbook for files
        files_sheet = workbook.active
        files_sheet.title = "Files"
        files_sheet.append(['File Name', 'Bilinguals', 'Sent'])
        for filename, existence in file_existence.items():
            if filename != 'PST Count':
                files_sheet.append([filename, existence[0], existence[1]])

        # Add auto-filter to columns B and C
        files_sheet.auto_filter.ref = files_sheet.dimensions

        # Resize columns to fit data
        for column_cells in files_sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            files_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Second workbook for PSTs
        pst_sheet = workbook.create_sheet(title="PST")
        pst_sheet.append(['PST Files'])
        for file in os.listdir(os.path.join(base_folder, "admin")):
            if file.lower().endswith('.pst'):
                pst_sheet.append([file])

        # Resize columns to fit data for PST sheet
        for column_cells in pst_sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            pst_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Third workbook for Quotes
        quote_sheet = workbook.create_sheet(title="Quotes")
        quote_sheet.append(['Quotes'])
        for file in os.listdir(os.path.join(base_folder, "admin", "Quote")):
            if file.lower().endswith('.xlsx'):
                quote_sheet.append([file])

        # Fourth workbook for Invoices
        invoices_sheet = workbook.create_sheet(title="Invoices")
        invoices_sheet.append(['Invoices'])
        invoices_folder = os.path.join(base_folder, "admin", "_Invoice")
        if not os.path.exists(invoices_folder):
            invoices_folder = os.path.join(base_folder, "admin", "Invoice")
        for file in os.listdir(invoices_folder):
            if file.lower().endswith('.pdf'):
                invoices_sheet.append([file])

        # Resize columns to fit data for Invoices sheet
        for column_cells in invoices_sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            invoices_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Fifth workbook for memoQ Archive
        archive_sheet = workbook.create_sheet(title="memoQ Archive")
        archive_sheet.append(['memoQ Archive'])
        archive_folder_name = os.path.basename(os.path.normpath(base_folder))
        for file in os.listdir(os.path.join("P:/_MMQGarbage/Archive")):
            if file.lower().endswith('.mqarch') and archive_folder_name in file:
                archive_sheet.append([file])

        # Resize columns to fit data for memoQ Archive sheet
        for column_cells in archive_sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            archive_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        try:
            workbook.save(xlsx_file_path)
            messagebox.showinfo("Export Complete", "Excel file exported successfully!")
        except PermissionError:
            messagebox.showerror("Error", "Permission denied! Please close the file and try again.")

def restart_process():
    global base_folder
    base_folder = filedialog.askdirectory(title="Select Base Folder")  # Ask for a new base folder
    if base_folder:
        root.title(f"Audit Check AVA - {os.path.basename(base_folder)}")  # Update window title
        clear_widgets(root)  # Clear existing widgets
        display_results(base_folder)

def clear_widgets(root):
    for widget in root.winfo_children():
        widget.destroy()

def show_progress_bar():
    global progress_window
    global total_files_count
    global processed_files_count

    progress_window = tk.Toplevel()
    progress_window.title("Progress")
    progress_window.geometry("300x120")  # Increased height to accommodate the label

    progress_label = tk.Label(progress_window, text="Analyzing folders...")
    progress_label.pack(pady=10)

    global file_count_label  # Declare file count label as global
    file_count_label = tk.Label(progress_window, text=f"Processed Files: {processed_files_count}/{total_files_count}")
    file_count_label.pack(pady=5)

    progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
    progress_bar.pack(pady=5)
    progress_bar.start()

def update_progress():
    global processed_files_count

    processed_files_count += 1
    progress_percentage = int((processed_files_count / total_files_count) * 100)
    progress_bar["value"] = progress_percentage
    progress_window.update_idletasks()

    # Update the file count label
    file_count_label.config(text=f"Processed Files: {processed_files_count}/{total_files_count}")

def display_results(base_folder):
    global progress_window

    def open_folder():
        # Convert the base_folder path to a format that Windows Explorer expects
        explorer_path = os.path.join(os.environ["WINDIR"], "explorer.exe")
        subprocess.Popen([explorer_path, os.path.normpath(base_folder)])

    show_progress_bar()  # Show the progress bar window

    root = tk.Tk()
    root.title(f"Audit Check AVA - {os.path.basename(base_folder)}")  # Update window title to include folder name
    # Increase window height by 100 pixels
    root.geometry("1000x800")  # Increased height to accommodate the lower frame

    # Create top line frame
    top_line_frame = tk.Frame(root)
    top_line_frame.pack(side=tk.TOP, fill=tk.X)

    # PST Count
    pst_files = os.listdir(os.path.join(base_folder, "admin"))
    pst_count = len([file for file in pst_files if file.endswith('.pst')])
    pst_text = f"PST Count: {pst_count}"
    pst_label = tk.Label(top_line_frame, text=pst_text)
    pst_label.pack(side=tk.LEFT, padx=10)
    pst_listbox = tk.Listbox(top_line_frame, height=5)
    for file in pst_files:
        if file.endswith('.pst'):
            pst_listbox.insert(tk.END, file)
    pst_listbox.pack(side=tk.LEFT, padx=10)

    # Quotes
    quote_folder = os.path.join(base_folder, "admin", "Quote")
    quote_files = os.listdir(quote_folder)
    quote_count = len(quote_files)
    quote_text = f"Quotes: {quote_count}"
    quote_label = tk.Label(top_line_frame, text=quote_text)
    quote_label.pack(side=tk.LEFT, padx=10)
    quote_listbox = tk.Listbox(top_line_frame, height=5)
    for file in quote_files:
        quote_listbox.insert(tk.END, file)
    quote_listbox.pack(side=tk.LEFT, padx=10)

    # Invoices
    # Handling both "_Invoice" and "Invoice" folder names
    invoice_folder = os.path.join(base_folder, "admin", "_Invoice").replace("\\", "/")
    if not os.path.exists(invoice_folder):
        invoice_folder = os.path.join(base_folder, "admin", "Invoice").replace("\\", "/")
    invoice_files = os.listdir(invoice_folder)
    invoice_count = len(invoice_files)
    invoice_text = f"Invoices: {invoice_count}"
    invoice_label = tk.Label(top_line_frame, text=invoice_text)
    invoice_label.pack(side=tk.LEFT, padx=10)
    invoice_listbox = tk.Listbox(top_line_frame, height=5)
    for file in invoice_files:
        invoice_listbox.insert(tk.END, file)
    invoice_listbox.pack(side=tk.LEFT, padx=10)

    # Reference Files
    reference_folder = os.path.join(base_folder, "reference")
    reference_files = os.listdir(reference_folder)
    reference_count = len(reference_files)
    reference_text = f"Reference Files: {reference_count}"
    reference_label = tk.Label(top_line_frame, text=reference_text)
    reference_label.pack(side=tk.LEFT, padx=10)
    reference_listbox = tk.Listbox(top_line_frame, height=5)
    for file in reference_files:
        reference_listbox.insert(tk.END, file)
    reference_listbox.pack(side=tk.LEFT, padx=10)

    # Create bottom line frame
    bottom_line_frame = tk.Frame(root)
    bottom_line_frame.pack(side=tk.TOP, fill=tk.X, pady=10)  # Added pady=10 for space

    # Create and add Export button with padding
    export_button = tk.Button(bottom_line_frame, text="Export to XLSX", command=export_to_xlsx)
    export_button.pack(side=tk.BOTTOM, padx=100, pady=10)  # Added pady=10 for space

    # Create button to select another folder
    select_folder_button = tk.Button(bottom_line_frame, text="Select Another Folder", command=restart_process)
    select_folder_button.pack(side=tk.BOTTOM, padx=100, pady=10)  # Added pady=10 for space

    # Create button to open the selected folder in Windows Explorer
    open_folder_button = tk.Button(bottom_line_frame, text="Open Folder", command=open_folder)
    open_folder_button.pack(side=tk.BOTTOM, padx=100, pady=10)  # Added pady=10 for space

    # Files List with Bilinguals and Sent
    files_label = tk.Label(bottom_line_frame, text="File List:")
    files_label.pack(side=tk.TOP, padx=10)

    files_tree = ttk.Treeview(bottom_line_frame)
    files_tree.pack(side=tk.TOP, padx=10)

    # Add columns to the Treeview
    files_tree["columns"] = ("Bilinguals", "Sent")
    files_tree.heading("#0", text="Files")
    files_tree.heading("Bilinguals", text="Bilinguals")
    files_tree.heading("Sent", text="Sent")

    # Configure column options to center the text
    files_tree.column("Bilinguals", anchor="center")  # Center the Bilinguals column
    files_tree.column("Sent", anchor="center")  # Center the Sent column

    # Configure tag for rows with 'X' in Bilinguals or Sent columns
    files_tree.tag_configure("red_bg", background="red")

    # Get the file existence dictionary
    file_existence = compare_folders(base_folder)

    # Populate the Treeview
    for filename, existence in file_existence.items():
        # Ignore files in specific folders
        if "queries" in filename.lower() or "reference" in filename.lower() or "admin" in filename.lower():
            continue

        # Ensure existence is a list or tuple
        if isinstance(existence, (list, tuple)):
            # Insert row with tags for red background if 'X' is present in Bilinguals or Sent
            tags = ()
            if '✘' in existence[0] or '✘' in existence[1]:
                tags = ("red_bg",)
            files_tree.insert("", tk.END, text=filename, values=(existence[0], existence[1]), tags=tags)
        else:
            files_tree.insert("", tk.END, text=filename, values=(existence,), tags=("red_bg",))

    # Move rows with red background to the top
    for item_id in files_tree.get_children():
        if "red_bg" in files_tree.item(item_id, "tags"):
            files_tree.move(item_id, "", "0")

    # Create a frame for the third pane at the bottom
    archive_frame = tk.Frame(root)
    archive_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    # Add a label for the third pane
    archive_label = tk.Label(archive_frame, text="memoQ Archive:")
    archive_label.pack(side=tk.TOP, padx=10)

    # Add a listbox to display files in Archive
    global archive_listbox  # Making archive_listbox global for export_to_xlsx function
    archive_listbox = tk.Listbox(archive_frame, height=10)
    archive_listbox.pack(side=tk.TOP, padx=10, pady=5, fill=tk.BOTH, expand=True)

    # Get files in P:/_MMQGarbage/Archive
    archive_files = os.listdir("P:/_MMQGarbage/Archive")
    # Filter files that contain base folder name
    base_folder_name = os.path.basename(os.path.normpath(base_folder))
    filtered_files = [file for file in archive_files if base_folder_name in file]
    # Insert filtered files into listbox
    for file in filtered_files:
        archive_listbox.insert(tk.END, file)

    progress_window.destroy()
    root.mainloop()

def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()
        sys.exit()

if __name__ == "__main__":
    root = tk.Tk()
    root.protocol("WM_DELETE_WINDOW", on_closing)  # Call on_closing when the window is closed
    root.withdraw()  # Hide the main window

    base_folder = filedialog.askdirectory(title="Select Base Folder")

    if base_folder:
        display_results(base_folder)
    else:
        print("No folder selected.")

    root.mainloop()

